import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from dateutil.relativedelta import relativedelta
from io import BytesIO

st.set_page_config(page_title="KPI Generator", layout="wide")
st.title("📊 KPI Metrics Generator")

st.sidebar.header("📂 Upload Files")

# --- File Upload ---
ar_file = st.sidebar.file_uploader("Upload AR Analysis File (.xlsx)", type=["xlsx"])
dest_file = st.sidebar.file_uploader("Upload Destination Excel Template (.xlsx)", type=["xlsx"])

# --- Manual Inputs ---
st.sidebar.header("📋 Manual Inputs")
visits_input = st.sidebar.text_input("Visits count (Simplibill Dashboard)", "")
ar_31_60 = st.sidebar.text_input("A/R (31–60 Days)", "")
ar_61_90 = st.sidebar.text_input("A/R (61–90 Days)", "")
days_2025_input = st.sidebar.text_input("Days elapsed in 2025 (default auto)", "")

process = st.sidebar.button("Generate KPIs")

if process and ar_file and dest_file:
    st.info("Processing data...")

    # --- Calculate days ---
    try:
        days_2025 = int(days_2025_input)
    except:
        days_2025 = (datetime.today() - datetime(2025, 1, 1)).days

    denial_resolution = "85%"

    # --- Load AR data efficiently ---
    chunks = pd.read_excel(ar_file, dtype=str, chunksize=50000)
    df = pd.concat(chunks, ignore_index=True)

    df['Visit Date'] = pd.to_datetime(df['Visit Date'], errors='coerce')

    num_cols = [
        'Charge','Expected','Primary Payment','Secondary Payment',
        'Tertiary Payment','Patient Payment','Balance'
    ]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # ====================================================
    # Slide3 - Current KPI Metrics
    # ====================================================
    charges = df['Charge'].sum()
    expected = df['Expected'].sum()
    payments = df[['Primary Payment','Secondary Payment','Tertiary Payment','Patient Payment']].sum().sum()
    submitted_charges = df[df['Visit Status'].str.lower() == 'claim created']['Charge'].sum()

    charges_submitted_pct_val = (submitted_charges / charges * 100) if charges else 0
    charges_submitted_pct = f"{(charges_submitted_pct_val):.2f}%"
    gcr = f"{(payments / charges * 100):.2f}%" if charges else "0.00%"
    ncr_val = (payments / expected * 100) if expected else 0
    ncr = f"{ncr_val:.2f}%"

    avg_daily_charges = charges / 365 if charges else 0
    days_in_ar = round(df['Balance'].sum() / avg_daily_charges) if avg_daily_charges else 0

    kpi_data = pd.DataFrame([
        ["Visits", visits_input],
        ["Charges", f"${charges:,.2f}"],
        ["Charges Submitted (%)", charges_submitted_pct],
        ["Payments", f"${payments:,.2f}"],
        ["Gross Collection Rate (%)", gcr],
        ["Net Collection Rate (%)", ncr],
        ["Days in AR (DAR)", days_in_ar],
        ["A/R (31–60 Days)", ar_31_60],
        ["A/R (61–90 Days)", ar_61_90],
        ["Denial vs Resolution", denial_resolution]
    ], columns=["Type", "Value"])

    # ====================================================
    # Slide4 - Quarterly KPI Metrics
    # ====================================================
    def get_quarter_label(date):
        if pd.isna(date): return None
        q = (date.month - 1) // 3 + 1
        return f"{date.year} Q{q}"

    df['Quarter'] = df['Visit Date'].apply(get_quarter_label)

    def get_quarter_dates(q_label):
        year, q = q_label.split()
        year = int(year)
        quarter = int(q[1])
        start_month = (quarter - 1) * 3 + 1
        start_date = datetime(year, start_month, 1)
        end_date = start_date + relativedelta(months=3) - pd.Timedelta(days=1)
        return (start_date, end_date)

    target_quarters = ["2024 Q1","2024 Q2","2024 Q3","2024 Q4","2025 Q1","2025 Q2","2025 Q3"]
    slide4_data = []

    for q in target_quarters:
        q_df = df[df['Quarter'] == q]
        if q_df.empty:
            slide4_data.append([q] + ["N/A"]*12)
            continue

        visits = len(q_df)  # total rows, not unique
        charges = q_df['Charge'].sum()
        expected = q_df['Expected'].sum()
        payments = q_df[['Primary Payment','Secondary Payment','Tertiary Payment','Patient Payment']].sum().sum()
        submitted_charges = q_df[q_df['Visit Status'].str.lower() == 'claim created']['Charge'].sum()

        charges_submitted_pct = f"{(submitted_charges/charges*100):.2f}%" if charges else "0.00%"
        gcr = f"{(payments/charges*100):.2f}%" if charges else "0.00%"
        ncr = f"{(payments/expected*100):.2f}%" if expected else "0.00%"

        start_date, end_date = get_quarter_dates(q)
        days_in_q = (end_date - start_date).days + 1
        avg_daily_charges = charges / days_in_q if charges else 0
        days_in_ar = round(q_df['Balance'].sum() / avg_daily_charges) if avg_daily_charges else 0

        billed_ar = q_df[q_df['Visit Status'].str.lower() == 'claim created']['Balance'].sum()
        unbilled_ar = q_df[q_df['Visit Status'].str.lower() != 'claim created']['Balance'].sum()
        total_ar = billed_ar + unbilled_ar
        billed_pct = f"{(billed_ar/total_ar*100):.2f}%" if total_ar else "0.00%"
        unbilled_pct = f"{(unbilled_ar/total_ar*100):.2f}%" if total_ar else "0.00%"

        slide4_data.append([
            q, visits, f"${charges:,.2f}", charges_submitted_pct, f"${payments:,.2f}",
            gcr, ncr, days_in_ar, f"${billed_ar:,.2f}", billed_pct,
            f"${unbilled_ar:,.2f}", unbilled_pct, denial_resolution
        ])

    slide4_df = pd.DataFrame(slide4_data, columns=[
        "Quarter","Visits","Charges","Charges Submitted (%)","Payments","Gross Collection Rate (%)",
        "Net Collection Rate (%)","Days in AR","Billed AR","Billed AR (%)",
        "Unbilled AR","Unbilled AR (%)","Denial vs Resolution (%)"
    ])

    # ====================================================
    # Slide5 - Yearly KPI Metrics
    # ====================================================
    slide5_data = []
    for year in [2023, 2024, 2025]:
        start_date = datetime(year,1,1)
        end_date = datetime(year+1,1,1) if year<2025 else datetime.today()
        days_in_year = 365 if year==2023 else 366 if year==2024 else days_2025

        df_y = df[(df['Visit Date']>=start_date) & (df['Visit Date']<end_date)].copy()
        if df_y.empty:
            slide5_data.append([year] + ["N/A"]*10)
            continue

        visits = len(df_y)
        charges = df_y['Charge'].sum()
        expected = df_y['Expected'].sum()
        payments = df_y[['Primary Payment','Secondary Payment','Tertiary Payment','Patient Payment']].sum().sum()
        submitted_charges = df_y[df_y['Visit Status'].str.lower() == 'claim created']['Charge'].sum()

        billed_ar = df_y[df_y['Visit Status'].str.lower() == 'claim created']['Balance'].sum()
        unbilled_ar = df_y[df_y['Visit Status'].str.lower() != 'claim created']['Balance'].sum()

        charges_submitted_pct = f"{(submitted_charges/charges*100):.2f}%" if charges else "0.00%"
        gcr = f"{(payments/charges*100):.2f}%" if charges else "0.00%"
        ncr = f"{(payments/expected*100):.2f}%" if expected else "0.00%"
        avg_daily_charges = charges/days_in_year if charges else 0
        days_in_ar = round((billed_ar+unbilled_ar)/avg_daily_charges) if avg_daily_charges else 0

        slide5_data.append([
            year, visits, f"${charges:,.2f}", charges_submitted_pct, f"${payments:,.2f}",
            gcr, ncr, days_in_ar, f"${billed_ar:,.2f}", f"${unbilled_ar:,.2f}", denial_resolution
        ])

    slide5_df = pd.DataFrame(slide5_data, columns=[
        "Year","Visits","Charges","Charges Submitted (%)","Payments","Gross Collection Rate (%)",
        "Net Collection Rate (%)","Days in AR","Billed AR","Unbilled AR","Denial vs Resolution (%)"
    ])

    # ====================================================
    # Display Tables
    # ====================================================
    st.subheader("Slide 3 – Current KPI Metrics")
    st.dataframe(kpi_data, use_container_width=True)

    st.subheader("Slide 4 – Quarterly KPI Metrics")
    st.dataframe(slide4_df, use_container_width=True)

    st.subheader("Slide 5 – Yearly KPI Metrics")
    st.dataframe(slide5_df, use_container_width=True)

    # ====================================================
    # Write to Excel
    # ====================================================
    dest_bytes = BytesIO(dest_file.read())
    wb = openpyxl.load_workbook(dest_bytes)

    def write_sheet(ws_name, columns, data):
        if ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row: cell.value = None
        else:
            ws = wb.create_sheet(ws_name)
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=2, column=col_idx, value=col_name).font = Font(bold=True)
        for row_idx, row in enumerate(data, start=3):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        return ws

    write_sheet("Slide3", list(kpi_data.columns), kpi_data.values.tolist())
    write_sheet("Slide4", list(slide4_df.columns), slide4_df.values.tolist())
    write_sheet("Slide5", list(slide5_df.columns), slide5_df.values.tolist())

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        label="📥 Download Updated Excel File",
        data=output,
        file_name="Updated_KPI_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
